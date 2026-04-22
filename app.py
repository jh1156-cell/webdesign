import math
from collections import deque
from pathlib import Path

import openpyxl
import streamlit as st

st.set_page_config(page_title="브랜드별 단가계산기", page_icon="🧮", layout="centered")

SHEET_NAME = "거래처DB"
VENDOR_START_ROW = 3
VENDOR_END_ROW = 63
DATA_FILE_CANDIDATES = [
    "업무리스트.xlsx",
    "업무리스트 (1).xlsx",
]
MIN_MARGIN = 30000

st.markdown(
    """
    <style>
    .final-price-card {
        background: linear-gradient(135deg, #111827 0%, #1f2937 100%);
        border-radius: 20px;
        padding: 24px 22px;
        margin: 10px 0 20px 0;
        box-shadow: 0 10px 28px rgba(0,0,0,0.18);
        border: 1px solid rgba(255,255,255,0.08);
    }
    .final-price-label {
        color: #d1d5db;
        font-size: 0.95rem;
        margin-bottom: 8px;
    }
    .final-price-value {
        color: white;
        font-size: 2.35rem;
        font-weight: 800;
        line-height: 1.1;
        margin-bottom: 8px;
    }
    .final-price-sub {
        color: #fca5a5;
        font-size: 0.98rem;
        font-weight: 700;
    }
    .section-chip {
        display: inline-block;
        background: #f3f4f6;
        color: #111827;
        border-radius: 999px;
        padding: 6px 12px;
        font-size: 0.88rem;
        font-weight: 700;
        margin-bottom: 10px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


def find_data_file() -> Path:
    base_dir = Path(__file__).parent
    for filename in DATA_FILE_CANDIDATES:
        candidate = base_dir / filename
        if candidate.exists():
            return candidate
    raise FileNotFoundError(
        "엑셀 파일을 찾지 못했습니다. 업무리스트.xlsx 또는 업무리스트 (1).xlsx 파일을 앱 폴더에 넣어주세요."
    )



def normalize_text(value):
    if value is None:
        return ""
    return "".join(str(value).split()).lower()


@st.cache_data
def load_vendor_data(file_path_str: str):
    file_path = Path(file_path_str)
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"'{SHEET_NAME}' 시트를 찾지 못했습니다.")

        ws = workbook[SHEET_NAME]
        rows = []
        for row in range(VENDOR_START_ROW, VENDOR_END_ROW + 1):
            vendor_name = ws[f"A{row}"].value
            rule_text = ws[f"F{row}"].value

            vendor_name = "" if vendor_name is None else str(vendor_name).strip()
            rule_text = "" if rule_text is None else str(rule_text).strip()

            if vendor_name:
                rows.append({"거래처": vendor_name, "기준정보": rule_text})
        return rows
    finally:
        workbook.close()



def round_up_to_10000(value):
    return int(math.ceil(value / 10000.0) * 10000)



def won(value):
    return f"{int(round(value)):,}원"



def calculate_values(x_value, c_value):
    purchase_price = x_value / 2
    markup_price = purchase_price * 1.4
    difference = c_value - purchase_price
    selling_price = c_value * 1.1
    return {
        "매입가": purchase_price,
        "1.4배": markup_price,
        "상품가": c_value,
        "차액": difference,
        "판매가": selling_price,
    }



def push_history(entry):
    history = st.session_state.setdefault("calc_history", deque(maxlen=5))
    history.appendleft(entry)



def render_final_price_card(selling_price: int, increase_amount: int):
    increase_html = (
        f"<div class='final-price-sub'>보정으로 판매가에 추가 반영: {won(increase_amount)}</div>"
        if increase_amount > 0
        else "<div class='final-price-sub' style='color:#86efac;'>추가 보정 없이 계산됨</div>"
    )
    st.markdown(
        f"""
        <div class="final-price-card">
            <div class="final-price-label">최종 결과</div>
            <div class="final-price-value">판매가(E) {won(selling_price)}</div>
            {increase_html}
        </div>
        """,
        unsafe_allow_html=True,
    )



def render_history():
    history = list(st.session_state.get("calc_history", []))
    st.markdown("### 최근 계산 결과")
    if not history:
        st.caption("아직 저장된 계산 결과가 없습니다.")
        return

    for idx, item in enumerate(history, start=1):
        with st.expander(f"{idx}. {item['거래처']} / 거래처가구가격={won(item['x'])}", expanded=(idx == 1)):
            st.write(f"**기준정보:** {item['기준정보']}")
            st.metric("최종 판매가(E)", won(item["판매가"]))
            if item["보정값"] > 0:
                st.markdown(
                    f"<p style='color:red; font-weight:700;'>차액(D) 보정액: {won(item['보정값'])}</p>",
                    unsafe_allow_html=True,
                )
                st.markdown(
                    f"<p style='color:red; font-weight:700;'>판매가(E) 추가 반영액: {won(item['판매가증가분'])}</p>",
                    unsafe_allow_html=True,
                )
            else:
                st.caption("차액 보정 없음")

            col1, col2 = st.columns(2)
            with col1:
                st.metric("매입가(A)", won(item["매입가"]))
                st.metric("1.4배(B)", won(item["1.4배"]))
                st.metric("상품가(C)", won(item["상품가"]))
            with col2:
                st.metric("차액(D)", won(item["차액"]))
                st.metric("판매가(E)", won(item["판매가"]))


if "saved_signature" not in st.session_state:
    st.session_state["saved_signature"] = None

try:
    data_file = find_data_file()
    vendor_rows = load_vendor_data(str(data_file))
except Exception as exc:
    st.error(f"데이터를 불러오는 중 오류가 발생했습니다: {exc}")
    st.stop()

if not vendor_rows:
    st.error("거래처 데이터를 찾지 못했습니다. 엑셀 범위(A3:A63, F3:F63)를 확인해 주세요.")
    st.stop()

st.title("브랜드별 단가계산기")
st.caption("거래처명을 일부 입력해서 선택하고, 계산 가능 여부를 바로 확인할 수 있습니다.")
st.caption(f"현재 데이터 파일: {data_file.name}")

search_text = st.text_input("거래처 검색", placeholder="예: FM, 꼬모도, 현대컴퍼니")
search_keyword = search_text.strip().lower()

if search_keyword:
    filtered_rows = [row for row in vendor_rows if search_keyword in row["거래처"].lower()]
else:
    filtered_rows = vendor_rows[:]

if not filtered_rows:
    st.warning("검색 결과가 없습니다. 다른 키워드로 다시 검색해 주세요.")
    render_history()
    st.stop()

selected_vendor = st.selectbox("거래처 선택", [row["거래처"] for row in filtered_rows], index=0)
selected_row = next(row for row in vendor_rows if row["거래처"] == selected_vendor)
rule_text = selected_row["기준정보"]
normalized_rule = normalize_text(rule_text)

st.markdown("### 선택된 거래처 정보")
st.write(f"**거래처명:** {selected_vendor}")
st.write(f"**F열 기준정보:** {rule_text if rule_text else '(빈칸)'}")

if "반값x1.4" not in normalized_rule:
    st.info("이 거래처는 자동 계산 대상이 아닙니다.")
    if rule_text:
        st.write(f"해당 기준정보: **{rule_text}**")
    else:
        st.write("해당 기준정보: **빈칸**")
    render_history()
    st.stop()

st.success("이 거래처는 `반값 x1.4` 규칙으로 계산할 수 있습니다.")

x_value = st.number_input("거래처가구가격", min_value=0, value=0, step=1000)

purchase_price = x_value / 2
markup_price = purchase_price * 1.4
base_product_price = round_up_to_10000(markup_price)
base_difference = base_product_price - purchase_price
adjustment_needed = max(0, MIN_MARGIN - base_difference)
recommended_c_value = base_product_price + adjustment_needed

st.markdown("<div class='section-chip'>기본 계산</div>", unsafe_allow_html=True)
col_base1, col_base2, col_base3 = st.columns(3)
col_base1.metric("매입가(A)", won(purchase_price))
col_base2.metric("1.4배(B)", won(markup_price))
col_base3.metric("기본 상품가(C)", won(base_product_price))

if adjustment_needed > 0:
    st.markdown(
        f"<p style='color:red; font-weight:700;'>차액(D)이 30,000원보다 {won(adjustment_needed)} 부족해서 그만큼 상품가(C)에 더했습니다.</p>",
        unsafe_allow_html=True,
    )
    c_value = st.number_input(
        "상품가(C) (수정 가능)",
        min_value=0,
        value=int(recommended_c_value),
        step=10000,
        help="보정값이 반영된 추천 상품가입니다. 필요하면 직접 수정할 수 있습니다.",
    )
else:
    st.caption("차액(D)이 이미 30,000원 이상이라 상품가(C) 수정이 잠겨 있습니다.")
    c_value = int(base_product_price)
    st.number_input(
        "상품가(C) (수정 불가)",
        min_value=0,
        value=int(base_product_price),
        step=10000,
        disabled=True,
    )

result = calculate_values(x_value, c_value)
actual_margin_gap = max(0, MIN_MARGIN - result["차액"])
added_to_c = max(0, c_value - base_product_price)
added_to_e = max(0, result["판매가"] - (base_product_price * 1.1))

st.markdown("### 판매가(E)")
render_final_price_card(int(result["판매가"]), int(added_to_e))

if actual_margin_gap > 0:
    st.markdown(
        f"<p style='color:red; font-weight:700;'>현재 상품가(C) 기준으로는 차액(D)이 아직 {won(actual_margin_gap)} 부족합니다.</p>",
        unsafe_allow_html=True,
    )
elif added_to_c > 0:
    st.markdown(
        f"<p style='color:red; font-weight:700;'>상품가(C)에 총 {won(added_to_c)} 추가 반영되었습니다.</p>",
        unsafe_allow_html=True,
    )

st.markdown("### 계산 상세")
col1, col2 = st.columns(2)
with col1:
    st.metric("매입가(A)", won(result["매입가"]))
    st.metric("1.4배(B)", won(result["1.4배"]))
    st.metric("상품가(C)", won(result["상품가"]))
with col2:
    st.metric("차액(D)", won(result["차액"]))
    st.metric("판매가(E)", won(result["판매가"]))

current_signature = (
    selected_vendor,
    rule_text,
    int(x_value),
    int(c_value),
    int(result["차액"]),
    int(added_to_c),
    int(result["판매가"]),
)

if st.button("현재 계산 결과 저장"):
    if st.session_state.get("saved_signature") != current_signature:
        push_history(
            {
                "거래처": selected_vendor,
                "기준정보": rule_text if rule_text else "(빈칸)",
                "x": int(x_value),
                "보정값": int(added_to_c),
                "판매가증가분": int(added_to_e),
                **{k: int(v) for k, v in result.items()},
            }
        )
        st.session_state["saved_signature"] = current_signature
        st.success("최근 계산 결과에 저장했습니다.")
    else:
        st.info("같은 계산 결과는 이미 저장되어 있습니다.")

render_history()

with st.expander("계산식 보기"):
    st.write("A = x / 2")
    st.write("B = (x / 2) × 1.4")
    st.write("C = B를 10,000원 단위 올림")
    st.write("D = C - A")
    st.write("D가 30,000원 미만이면 부족한 금액만큼 C를 올림")
    st.write("E = C × 1.1")
